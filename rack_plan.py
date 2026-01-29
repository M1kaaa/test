"""
Загрузка и работа с планом стоек из Excel.

Файл: racks-condition (10).xlsx

Ограничиваемся стойками с кодами 02b03–02b18 (включительно).

Так как структура файла может отличаться, модуль старается:
- автоматически найти колонку с названием стойки (заголовок содержит 'rack');
- взять активный лист;
- построить отображение: человекочитаемый код стойки -> порядковый индекс
  вдоль ряда (1, 2, 3, ...) только для нужного диапазона стоек.

При необходимости можно скорректировать функцию _detect_rack_column().
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


EXCEL_FILE_NAME = "racks-condition (10).xlsx"

# Диапазон интересующих стоек
RACK_CODE_START = "02b03"
RACK_CODE_END = "02b18"


@dataclass(frozen=True)
class RackInfo:
    """Информация о стойке, доступная во фронтенде и API."""

    code: str  # исходный код из Excel, например "02b05"
    index: int  # порядковый индекс в ряду (1, 2, 3, ...)


def _detect_rack_column(sheet) -> Optional[int]:
    """
    Попытаться автоматически определить индекс колонки с именем стойки.

    Стратегия:
    - просматриваем первые 10 строк;
    - ищем ячейку, где в тексте встречается подстрока 'rack' (без регистра);
    - возвращаем номер колонки (1-based) или None.
    """
    max_header_rows = 10

    for row in sheet.iter_rows(min_row=1, max_row=max_header_rows):
        for cell in row:
            value = str(cell.value).strip() if cell.value is not None else ""
            if value and "rack" in value.lower():
                return cell.column

    return None


def _parse_rack_code(code: str) -> Tuple[str, int]:
    """
    Разобрать код стойки вида '02b05' на (префикс, число).

    Предполагаем формат: <зона><буква><две цифры>, например '02b05'.
    Возвращаем:
        prefix = '02b'
        number = 5
    """
    code = code.strip()
    if len(code) < 4:
        return code, 0

    prefix = code[:-2]
    try:
        number = int(code[-2:])
    except ValueError:
        number = 0

    return prefix, number


def _is_in_interesting_range(code: str) -> bool:
    """Проверить, попадает ли код стойки в диапазон 02b03–02b18."""
    if not code:
        return False

    # Нормализуем к нижнему регистру
    code = code.strip()
    start = RACK_CODE_START.strip()
    end = RACK_CODE_END.strip()

    # Быстрая проверка по префиксу
    p_code, n_code = _parse_rack_code(code)
    p_start, n_start = _parse_rack_code(start)
    p_end, n_end = _parse_rack_code(end)

    if p_code.lower() != p_start.lower() or p_code.lower() != p_end.lower():
        return False

    return n_start <= n_code <= n_end


def generate_default_rack_plan() -> Tuple[List[RackInfo], Dict[str, RackInfo]]:
    """
    Сгенерировать дефолтный план стоек 02b03–02b18 без чтения Excel.

    Используется как резервный вариант, если файл Excel отсутствует
    или его структура неожиданная.
    """
    prefix, start_num = _parse_rack_code(RACK_CODE_START)
    _, end_num = _parse_rack_code(RACK_CODE_END)

    rack_infos: List[RackInfo] = []
    code_to_info: Dict[str, RackInfo] = {}

    idx = 1
    for n in range(start_num, end_num + 1):
        code = f"{prefix}{n:02d}"
        info = RackInfo(code=code, index=idx)
        rack_infos.append(info)
        code_to_info[code] = info
        idx += 1

    return rack_infos, code_to_info


def load_rack_plan(
    base_dir: Path | str | None = None,
) -> Tuple[List[RackInfo], Dict[str, RackInfo]]:
    """
    Загрузить план стоек и вернуть:
    - список RackInfo (отсортирован по индексу);
    - словарь code -> RackInfo.

    Если файл или нужная колонка не найдены, будет выброшено исключение.
    """
    if base_dir is None:
        base_dir = Path(__file__).resolve().parent
    else:
        base_dir = Path(base_dir)

    excel_path = base_dir / EXCEL_FILE_NAME
    if not excel_path.exists():
        raise FileNotFoundError(
            f"Файл плана стоек не найден: {excel_path}. "
            f"Скорректируйте EXCEL_FILE_NAME в rack_plan.py."
        )

    wb = load_workbook(excel_path, data_only=True)
    sheet = wb.active

    rack_col = _detect_rack_column(sheet)
    if rack_col is None:
        raise RuntimeError(
            "Не удалось автоматически определить колонку с кодом стойки. "
            "Проверьте заголовки в Excel и при необходимости "
            "обновите функцию _detect_rack_column()."
        )

    # Собираем все коды стоек
    rack_codes: List[str] = []
    header_row_detected = False
    header_row_index: Optional[int] = None

    # Сначала найдём строку заголовка (там, где находится rack_col)
    max_header_rows = 10
    for row in sheet.iter_rows(min_row=1, max_row=max_header_rows):
        cell = row[rack_col - 1]
        value = str(cell.value).strip() if cell.value is not None else ""
        if value and "rack" in value.lower():
            header_row_detected = True
            header_row_index = cell.row
            break

    if not header_row_detected or header_row_index is None:
        # fallback: считаем, что заголовок в первой строке
        header_row_index = 1

    # Теперь читаем все значения ниже заголовка
    for row in sheet.iter_rows(min_row=header_row_index + 1):
        cell = row[rack_col - 1]
        if cell.value is None:
            continue
        code = str(cell.value).strip()
        if not code:
            continue
        if _is_in_interesting_range(code):
            rack_codes.append(code)

    # Удаляем дубликаты и сортируем по числовой части
    unique_codes = sorted(set(rack_codes), key=lambda c: _parse_rack_code(c)[1])

    rack_infos: List[RackInfo] = []
    code_to_info: Dict[str, RackInfo] = {}

    for idx, code in enumerate(unique_codes, start=1):
        info = RackInfo(code=code, index=idx)
        rack_infos.append(info)
        code_to_info[code] = info

    if not rack_infos:
        raise RuntimeError(
            f"Не найдено ни одной стойки в диапазоне {RACK_CODE_START}–{RACK_CODE_END}. "
            "Проверьте содержимое Excel-файла."
        )

    return rack_infos, code_to_info


def get_rack_index_by_code(
    code: str,
    code_to_info: Dict[str, RackInfo],
) -> int:
    """
    Получить порядковый индекс стойки по её коду.

    Если код не найден, выбрасывается KeyError.
    """
    code = code.strip()
    info = code_to_info.get(code)
    if info is None:
        raise KeyError(f"Стойка с кодом '{code}' не найдена в загруженном плане.")
    return info.index

